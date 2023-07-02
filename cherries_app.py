from tkinter import *
from PIL import ImageTk, Image
import sqlite3
import csv
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import pandas as pd
from tkinter import ttk
import locale


if __name__ == "__main__":
    root = Tk()
    root.title("Cherries App")
    root.geometry("720x740")
    root.config(background="#EEE8D1")
    icon = PhotoImage(file="kerasia.png")
    root.iconphoto(True, icon)



    title_label = Label(root, text="    ΚΕΡΑΣΙΑ DATABASE   ", font=("arial black", 20), background="black",
                        relief="ridge", borderwidth=6, foreground="red")
    title_label.grid(row=0, column=0, columnspan=2, pady=20, padx=20)

    conn = sqlite3.connect("cherries_file.db")
    c = conn.cursor()

    c.execute("""CREATE TABLE IF NOT EXISTS products (
                date integer,
                variety text,
                items integer,
                kilo_m text,
                kilo_k text,
                price float)""")

    variety_combobox = ttk.Combobox(root, font=("arial black", 13), width=15)
    variety_combobox.grid(row=9, columnspan=2, pady=10)

    def update_variety_list():
        global variety_combobox
        c.execute("SELECT DISTINCT variety FROM products")
        variety = [row[0] for row in c.fetchall()]
        variety_combobox['values'] = variety



    def update():
        conn = sqlite3.connect("cherries_file.db")
        c = conn.cursor()
        show_id = delete_box.get()
        c.execute("""UPDATE products SET
               date = :date,
               variety = :variety,
               items = :items,
               kilo_m = :kilo_m,
               kilo_k = :kilo_k,
               price = :price
    
    
               WHERE oid = :oid""",
                  {'date': date_editor.get(),
                   'variety': variety_editor.get(),
                   'items': items_editor.get(),
                   'kilo_m': kilo_m_editor.get(),
                   'kilo_k': kilo_k_editor.get(),
                   'price': price_editor.get(),
                   'oid': show_id})

        confirmation = messagebox.askquestion("Επεξεργασία Αρχείων", "Θέλετε να αποθηκεύσετε τις αλλαγές  ???")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo(" Αποθήκευση Αρχείων ", "Επιτυχής αλλαγή και αποθήκευση Αρχείων. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Εισαγωγής", "Η αλλαγή Αρχείων ακυρώθηκε. !!!")

        conn.close()
        editor.destroy()

    def save_to_excel():
        conn = sqlite3.connect("cherries_file.db")
        c = conn.cursor()

        c.execute("SELECT * FROM products")
        data = c.fetchall()

        try:
            workbook = load_workbook("total.cherries.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            headers = ["ΗΜΕΡΟΜΗΝΙΑ", "ΠΟΙΚΙΛΙΑ", "ΤΕΛΑΡΑ", "ΚΙΛΑ ΜΕΙΚΤΑ", "ΚΙΛΑ ΚΑΘΑΡΑ", "ΤΙΜΗ"]
            sheet.append(headers)

        # Clear existing data in the worksheet
        sheet.delete_rows(2, sheet.max_row)

        for row in data:
            sheet.append(row)

        workbook.save("total.cherries.xlsx")

        conn.close()

        messagebox.showinfo("Αποθήκευση Δεδομένων", "Τα δεδομένα αποθηκεύτηκαν σε αρχείο Excel !!!")


    def submit():
        conn = sqlite3.connect("cherries_file.db")
        c = conn.cursor()

        price_value = price.get().replace(',', '.')
        if price_value == "0" or not price_value.strip():
            price_value = ""

        c.execute("INSERT INTO products VALUES (:date, :variety, :items, :kilo_m, :kilo_k, :price)",
                  {
                      "date": date.get(),
                      "variety": variety.get(),
                      "items": items.get() + "  ΤΕΛΑΡΑ  ",
                      "kilo_m": kilo_m.get() + "  ΚΙΛΑ ΜΕΙΚΤΑ  ",
                      "kilo_k": kilo_k.get() + "  ΚΙΛΑ ΚΑΘΑΡΑ  ",
                      "price": price.get()

                      })






        confirmation = messagebox.askquestion("Εισαγωγή Αρχείων", "θέλετε να εισάγετε αυτά τα Αρχεία ;")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo(" Εισαγωγή Αρχείων ", "Επιτυχής Εισαγωγή Αρχείων. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Εισαγωγής", "Η Εισαγωγή Αρχείων ακυρώθηκε. !!!")
        conn.close()

        date.delete(0, END)
        variety.delete(0, END)
        items.delete(0, END)
        kilo_m.delete(0, END)
        kilo_k.delete(0, END)
        price.delete(0, END)

    def delete_all_data():
        confirmation = messagebox.askyesno("Confirmation", "Είστε σίγουρος ότι θέλετε να διαγράψετε όλα τα Αρχεία ;")
        if confirmation:
            conn = sqlite3.connect("cherries_file.db")
            c = conn.cursor()
            c.execute("DELETE FROM products")
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Ολα τα Αρχεία διαγράφηκαν με επιτυχία !!")


    def show():
        show = Tk()
        show.title("ΚΕΡΑΣΙΑ DATABASE")
        show.geometry("800x800")
        show.config(background="#E9E8D4")



        show_id = delete_box.get()
        conn = sqlite3.connect("cherries_file.db")
        c = conn.cursor()

        c.execute("SELECT *,  oid FROM products")
        elements = c.fetchall()

        for i, element in enumerate(elements):
            text = '   ' .join(str(item) for item in element)
            show_label = Label(show, text=text, font=("arial black", 10), relief="ridge", bd=7, borderwidth=5, background="#D1F3F3", foreground="#20047A")
            show_label.grid(row=i, column=0, pady=1, sticky=W, ipadx=30)

        save_button = Button(show, text="Αποθήκευση σε 'Excel' ", command=save_to_excel, font=("arial black", 11), background="#06D784")
        save_button.grid(row=i + 1, column=0, pady=10, ipadx=26, sticky=W, padx=30)
        delete_button = Button(show, text="Διαγραφή όλων των Αρχείων", command=delete_all_data, font=("arial black", 11),
                               background="#EA6969")
        delete_button.grid(row=i + 1, column=0, pady=10, padx=450, ipadx=8)



    def delete():
        conn = sqlite3.connect("cherries_file.db")
        c = conn.cursor()

        c.execute("DELETE from products WHERE oid = " + delete_box.get())

        confirmation = messagebox.askquestion("Διαγραφή Στοιχείων", " Θέλετε να διαγράψετε τα Στοιχεία ;")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo("Η Διαγραφή Ολοκληρώθηκε", "Τα στοιχεία διαγράφηκαν με επιτυχία. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Διαγραφής", "Η διαγραφή ακυρώθηκε. !!!")

        conn.commit()
        conn.close()

    def edit():
        global editor
        editor = Tk()
        editor.title("ΕΠΕΞΕΡΓΑΣΙΑ ΣΤΟΙΧΕΙΩΝ")
        editor.geometry("500x400")
        editor.config(background="#EEE8D1")
        icon = PhotoImage(file="kerasia.png")
        root.iconphoto(True, icon)

        conn = sqlite3.connect("cherries_file.db")
        c = conn.cursor()

        show_id = delete_box.get()

        c.execute("SELECT * FROM products WHERE oid = " + show_id)
        shows = c.fetchall()

        global date_editor
        global variety_editor
        global items_editor
        global kilo_m_editor
        global kilo_k_editor
        global price_editor


        date_editor = Entry(editor, width=20, font=("arial black", 12), bd=2)
        date_editor.grid(row=0, column=1, padx=30, pady=(10, 0))
        variety_editor = Entry(editor, width=20, font=("arial black", 12), bd=2)
        variety_editor.grid(row=1, column=1, padx=30)
        items_editor = Entry(editor, width=20, font=("arial black", 12), bd=2)
        items_editor.grid(row=2, column=1, padx=30)
        kilo_m_editor = Entry(editor, width=20, font=("arial black", 12), bd=2)
        kilo_m_editor.grid(row=3, column=1, padx=30)
        kilo_k_editor = Entry(editor, width=20, font=("arial black", 12), bd=2)
        kilo_k_editor.grid(row=4, column=1, padx=30)
        price_editor = Entry(editor, width=20, font=("arial black", 12), bd=2)
        price_editor.grid(row=5, column=1, padx=30)


        date_editor_label = Label(editor, text="Ημερομηνία", font=("arial black", 12), background="#EEE8D1")
        date_editor_label.grid(row=0, column=0, pady=(10, 0), sticky=W, padx=10)
        variety_editor_label = Label(editor, text="Ποικιλία", font=("arial black", 12), background="#EEE8D1")
        variety_editor_label.grid(row=1, column=0, sticky=W, padx=10)
        items_editor_label = Label(editor, text="Τελάρα", font=("arial black", 12), background="#EEE8D1")
        items_editor_label.grid(row=2, column=0, sticky=W, padx=10)
        kilo_m_editor_label = Label(editor, text="Κιλά Μεικτά", font=("arial black", 12), background="#EEE8D1")
        kilo_m_editor_label.grid(row=3, column=0, sticky=W, padx=10)
        kilo_k_editor_label = Label(editor, text="Κιλά Καθαρά", font=("arial black", 12), background="#EEE8D1")
        kilo_k_editor_label.grid(row=4, column=0, sticky=W, padx=10)
        price_editor_label = Label(editor, text="Τιμή", font=("arial black", 12), background="#EEE8D1")
        price_editor_label.grid(row=5, column=0, sticky=W, padx=10)


        for show in shows:
            date_editor.insert(0, show[0])
            variety_editor.insert(0, show[1])
            items_editor.insert(0, show[2])
            kilo_m_editor.insert(0, show[3])
            kilo_k_editor.insert(0, show[4])
            price_editor.insert(0, show[5])


        edit_btn = Button(editor, text="Επεξεργασία και Αποθήκευση \n Στοιχείων", bd=4, command=update, font=("arial black", 10), background="#06D784", activeforeground="#06D784", activebackground="#06D784")
        edit_btn.grid(row=7, column=0, columnspan=2, pady=30, padx=10, ipadx=30)


    def open_total():
        total_w = Tk()
        total_w.title("ΣΥΝΟΛΑ")
        total_w.geometry("810x550")
        total_w.config(background="#EEE8D1")



        def calculate_total_euro():
            conn = sqlite3.connect("cherries_file.db")
            c = conn.cursor()


            c.execute("SELECT kilo_k, price FROM products")
            rows = c.fetchall()
            total_euro = sum(
                float(''.join(filter(str.isdigit, row[0]))) * float(str(row[1]).replace(',', '.')) for row in rows if
                row[0] and row[1])

            conn.close()
            total_euro_label.config(text="ΣΥΝΟΛΙΚΑ ΕΥΡΩ: {:.2f} €".format(total_euro))

            return total_euro


        def calculate_total_items():
            conn = sqlite3.connect("cherries_file.db")
            c = conn.cursor()
            c.execute("SELECT SUM(items) FROM products")
            total = c.fetchone()[0]
            if total is None:
                total = 0
            total_items_label.config(text="ΣΥΝΟΛΙΚΑ ΤΕΛΑΡΑ  : {} ".format(int(total)))

        def calculate_total_kilo_m():
            conn = sqlite3.connect("cherries_file.db")
            c = conn.cursor()
            c.execute("SELECT SUM(kilo_m) FROM products")
            total = c.fetchone()[0]
            if total is None:
                total = 0
            total_kilo_m_label.config(text="ΣΥΝΟΛΙΚΑ ΜΕΙΚΤΑ ΚΙΛΑ  : {} ".format(total))



        def calculate_total_kilo_k():
            conn = sqlite3.connect("cherries_file.db")
            c = conn.cursor()
            c.execute("SELECT SUM(kilo_k) FROM products")
            total = c.fetchone()[0]
            if total is None:
                total = 0
            total_kilo_k_label.config(text="ΣΥΝΟΛΙΚΑ ΚΑΘΑΡΑ ΚΙΛΑ  : {} ".format(total))

        def calculate_average_price():
            conn = sqlite3.connect("cherries_file.db")
            c = conn.cursor()

            c.execute("SELECT price FROM products")
            price = c.fetchall()



            valid_price = [float(str(price[0]).replace(',', '.')) for price in price if price[0]]
            total_price = sum(valid_price)
            average_price = total_price / len(valid_price) if valid_price else 0

            num_valid_prices = len(valid_price)
            average_price = total_price / num_valid_prices if num_valid_prices > 0 else 0

            average_price_label.config(text="ΜΕΣΗ ΤΙΜΗ: {:.2f} €".format(average_price))

            conn.commit()
            conn.close()

        def total_profit():
            profit_window = Tk()
            profit_window.title("Cherries App")
            profit_window.geometry("570x350")
            profit_window.config(background="#EEE8D1")

            workers = Entry(profit_window, width=15, font=("arial black", 14), bd=2)
            workers.grid(row=0, column=1, padx=20, pady=(20, 5))
            pesticide = Entry(profit_window, width=15, font=("arial black", 14), bd=2)
            pesticide.grid(row=1, column=1, padx=20, pady=5)

            workers_label = Label(profit_window, text="ΚΟΣΤΟΣ ΕΡΓΑΤΩΝ :", font=("arial black", 12), foreground="#721E8D",
                                  background="#EEE8D1")
            workers_label.grid(row=0, column=0, pady=(20, 0), sticky=W, padx=10)
            pesticide_label = Label(profit_window, text="ΦΥΤΟΦΑΡΜΑΚΑ & ΛΙΠΑΣΜΑΤΑ :", font=("arial black", 12), foreground="#721E8D",
                                    background="#EEE8D1")
            pesticide_label.grid(row=1, column=0, sticky=W, padx=10)

            def total_profit1(total_profit1_label):
                total_euro = calculate_total_euro()

                workers_value = float(workers.get())
                pesticide_value = float(pesticide.get().replace(",", "."))

                # Calculate the total.app profit
                total_profit = total_euro - workers_value - pesticide_value

                # Update the label with the new total.app profit value
                total_profit1_label.config(text="ΣΥΝΟΛΙΚΟ ΚΕΡΔΟΣ: {:.2f} €".format(total_profit))

            total_profit1_btn = Button(profit_window, text="ΣΥΝΟΛΙΚΟ ΚΕΡΔΟΣ:", font=("arial black", 10),
                                       command=lambda: total_profit1(total_profit1_label), bd=7,
                                       background="#06D784", activebackground="#06D784", activeforeground="#06D784")

            total_profit1_btn.grid(row=3, padx=5, columnspan=2, pady=(50, 20), ipadx=15)

            total_profit1_label = Label(profit_window, text="ΣΥΝΟΛΙΚΟ ΚΕΡΔΟΣ  :  0 € ", font=("arial black", 14), width=30,
                                        background="#A2D8F0", relief="ridge", borderwidth=10, foreground="#42047A")
            total_profit1_label.grid(row=4, columnspan=2, padx=10, pady=10, ipadx=15)


        calculate_euro_btn = Button(total_w, text="ΥΠΟΛΟΓΙΣΜΟΣ ΕΥΡΩ", command=calculate_total_euro, font=("arial black", 11),
                                    bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        calculate_euro_btn.grid(row=11, column=0, pady=(25, 5), padx=10, ipadx=77)

        total_euro_label = Label(total_w, text="ΣΥΝΟΛΟ ΕΥΡΩ   :  0 €   ", font=("arial black", 13), background="#A2D8F0",
                                 relief="ridge", borderwidth=5, foreground="#42047A", width=25)
        total_euro_label.grid(row=11, column=1, sticky=W, padx=10, pady=(25, 5), ipadx=50)

        calculate_items_btn = Button(total_w, text="ΥΠΟΛΟΓΙΣΜΟΣ ΤΕΛΑΡΑ", command=calculate_total_items, bd=4,
                                     font=("arial black", 11), background="#DE9C30", activebackground="#DE9C30",
                                     activeforeground="#DE9C30")
        calculate_items_btn.grid(row=12, column=0, pady=5, padx=10, ipadx=65)

        total_items_label = Label(total_w, text="ΣΥΝΟΛΙΚΑ ΤΕΛΑΡΑ  :  0  ", font=("arial black", 13), width=25,
                                  background="#A2D8F0", relief="ridge", borderwidth=5, foreground="#42047A")
        total_items_label.grid(row=12, column=1, sticky=W, padx=10, pady=5, ipadx=50)

        calculate_kilo_m_btn = Button(total_w, text="ΥΠΟΛΟΓΙΣΜΟΣ ΜΕΙΚΤΩΝ ΚΙΛΩΝ", command=calculate_total_kilo_m,
                                      font=("arial black", 11), bd=4, background="#DE9C30", activebackground="#DE9C30",
                                      activeforeground="#DE9C30")
        calculate_kilo_m_btn.grid(row=13, column=0, pady=5, padx=10, ipadx=28)

        total_kilo_m_label = Label(total_w, text="ΣΥΝΟΛΙΚΑ ΜΕΙΚΤΑ ΚΙΛΑ   :  0      ", font=("arial black", 13), width=25,
                                   background="#A2D8F0", relief="ridge", borderwidth=5, foreground="#42047A")
        total_kilo_m_label.grid(row=13, column=1, sticky=W, padx=10, pady=5, ipadx=50)

        calculate_kilo_k_btn = Button(total_w, text="ΥΠΟΛΟΓΙΣΜΟΣ ΚΑΘΑΡΩΝ ΚΙΛΩΝ", command=calculate_total_kilo_k,
                                      font=("arial black", 11), bd=4, background="#DE9C30", activebackground="#DE9C30",
                                      activeforeground="#DE9C30")
        calculate_kilo_k_btn.grid(row=14, column=0, pady=5, padx=10, ipadx=25)

        total_kilo_k_label = Label(total_w, text="ΣΥΝΟΛΙΚΑ ΚΑΘΑΡΑ ΚΙΛΑ   :  0      ", font=("arial black", 13), width=25,
                                   background="#A2D8F0", relief="ridge", borderwidth=5, foreground="#42047A")
        total_kilo_k_label.grid(row=14, column=1, sticky=W, padx=10, pady=5, ipadx=50)

        average_price_btn = Button(total_w, text="ΥΠΟΛΟΓΙΣΜΟΣ ΜΕΣΗΣ ΤΙΜΗΣ", command=calculate_average_price,
                                   font=("arial black", 11), bd=4, background="#DE9C30", activebackground="#DE9C30",
                                   activeforeground="#DE9C30")
        average_price_btn.grid(row=15, column=0, pady=5, padx=10, ipadx=40)

        average_price_label = Label(total_w, text="ΜΕΣΗ ΤΙΜΗ   :  0      ", font=("arial black", 13), width=25,
                                    background="#A2D8F0", relief="ridge", borderwidth=5, foreground="#42047A")
        average_price_label.grid(row=15, column=1, sticky=W, padx=10, pady=5, ipadx=50)

        total_profit_btn = Button(total_w, text="ΣΥΝΟΛΙΚΟ ΚΕΡΔΟΣ", font=("arial black", 11), command=total_profit, bd=4,
                                  background="#06D784", activebackground="#06D784", activeforeground="#06D784")
        total_profit_btn.grid(row=16, columnspan=2, pady=30, ipadx=55)


        title_label = Label(total_w, text=" ΚΕΡΑΣΙΑ DATABASE ", font=("arial black", 18), background="black",
                        relief="ridge", borderwidth=10, foreground="red")
        title_label.grid(row=0, pady=10, padx=10, columnspan=2)


    def select_variety():
        selected_variety = variety_combobox.get()
        if selected_variety:
            c.execute("SELECT variety, SUM(items), SUM(kilo_m), SUM(kilo_k), SUM(kilo_k * price) AS total_euro FROM products WHERE variety=? GROUP BY variety", (selected_variety,))
            variety_info = c.fetchall()
            price_value = price.get().replace(',', '.')

            if price_value == "0" or not price_value.strip():
                price_value = ""



            if variety_info:
                variety_window = Toplevel(root)
                variety_window.title("variety Information")
                variety_window.geometry("500x300")
                variety_window.config(background="#EEE8D1")


                info_labels = [" ΠΟΙΚΙΛΙΑ : ", " ΤΕΛΑΡΑ : ", " ΚΙΛΑ ΜΕΙΚΤΑ : ", " ΚΙΛΑ ΚΑΘΑΡΑ : ", "ΕΥΡΩ :"]

                for row_index, variety_data in enumerate(variety_info):
                    for i, label in enumerate(info_labels):
                        label_text = Label(variety_window, text=label, font=("arial black", 12), width=15, relief="ridge", bd=3,  background="#F1B52D")
                        label_text.grid(row=row_index * len(info_labels) + i, column=0, padx=30, sticky=W, pady=10)

                        if i == 4:
                            # Display the sum of prices
                            value_label = Label(variety_window, text=variety_data[i], font=("arial black", 13),  width=15,
                                  background="#A2D8F0", relief="ridge", borderwidth=5, foreground="#42047A")
                        else:
                            # Display the worker name
                            value_label = Label(variety_window, text=variety_data[i], font=("arial black", 13),  width=15,
                                  background="#A2D8F0", relief="ridge", borderwidth=5, foreground="#42047A")

                        value_label.grid(row=row_index * len(info_labels) + i, column=1, padx=30, pady=10)

            else:
                messagebox.showerror("Error", "Δεν υπάρχουν πληροφορίες για τη συκεκριμένη ποικιλία.")
        else:
            messagebox.showwarning("Warning", "Παρακαλώ επιλέξτε ποικιλία.")

        update_variety_list()

    update_variety_list()

    date = Entry(root, width=15, font=("arial black", 14), bd=3, background="#DBE5E5")
    date.grid(row=1, column=1, padx=50, pady=(20, 5))
    variety = Entry(root, width=15, font=("arial black", 14), bd=2, background="#DBE5E5")
    variety.grid(row=2, column=1, padx=50, pady=5)
    items = Entry(root, width=15, font=("arial black", 14), bd=2, background="#DBE5E5")
    items.grid(row=3, column=1, padx=50, pady=5)
    kilo_m = Entry(root, width=15, font=("arial black", 14), bd=2, background="#DBE5E5")
    kilo_m.grid(row=4, column=1, padx=50, pady=5)
    kilo_k = Entry(root, width=15, font=("arial black", 14), bd=2, background="#DBE5E5")
    kilo_k.grid(row=5, column=1, padx=50, pady=5)
    price = Entry(root, width=15, font=("arial black", 14), bd=2, background="#DBE5E5")
    price.grid(row=6, column=1, padx=50, pady=5)
    delete_box = Entry(root, width=6, font=("arial black", 14), bd=3, background="#DBE5E5")
    delete_box.grid(row=8, column=1, padx=20)


    date_label = Label(root, text="  Ημερομηνία : ", font=("arial black", 13),   background="#EEE8D1", foreground="#5D046B")
    date_label.grid(row=1, column=0, pady=(20, 0),sticky=W, padx=80)
    variety_label = Label(root, text="  Ποικιλία : ", font=("arial black", 13),  background="#EEE8D1", foreground="#5D046B")
    variety_label.grid(row=2, column=0, sticky=W, padx=80)
    items_label = Label(root, text="  Τελάρα : ", font=("arial black", 13), background="#EEE8D1", foreground="#5D046B")
    items_label.grid(row=3, column=0, sticky=W, padx=80)
    kilo_m_label = Label(root, text="  Κιλά Μεικτά : ", font=("arial black", 13),  background="#EEE8D1", foreground="#5D046B")
    kilo_m_label.grid(row=4, column=0,sticky=W, padx=80)
    kilo_k_label = Label(root, text="  Κιλά Καθαρά : ", font=("arial black", 13),  background="#EEE8D1", foreground="#5D046B")
    kilo_k_label.grid(row=5, column=0,sticky=W, padx=80)
    price_label = Label(root, text="  Τιμή : ", font=("arial black", 13), background="#EEE8D1", foreground="#5D046B")
    price_label.grid(row=6, column=0,sticky=W, padx=80)
    delete_box_label = Label(root, text="  Επιλογή ID ", font=("arial black", 14), background="#EEE8D1", foreground="#050EA3")
    delete_box_label.grid(row=8, column=0, sticky=W, padx=80, pady=13)


    submit_btn = Button(root, text="Εισαγωγή Στοιχείων", command=submit, font=("arial black", 12), bd=4, background="#06D784", activebackground="#06D784", activeforeground="#06D784")
    submit_btn.grid(row=10, column=0, pady=(20, 10), padx=10, ipadx=35)

    show_btn = Button(root, text="Εμφάνιση Στοιχείων", command=show, font=("arial black", 12), bd=4,  background="#0499A0", activebackground="#0499A0", activeforeground="#0499A0")
    show_btn.grid(row=11, column=1,  pady=10, padx=10, ipadx=35)

    delete_btn = Button(root, text="Διαγραφή Στοιχείων", command=delete, font=("arial black", 12), bd=4, background="#F67EA3", activebackground="#F67EA3", activeforeground="#F67EA3")
    delete_btn.grid(row=11, column=0, pady=10, padx=10, ipadx=35)

    edit_btn = Button(root, text="Επεξεργασία Στοιχείων", command=edit, font=("arial black", 12), bd=4, background="#A481C9", activebackground="#A481C9", activeforeground="#A481C9")
    edit_btn.grid(row=10, column=1, pady=(20, 10), padx=10, ipadx=23)





    total_btn = Button(root, text=" ΣΥΝΟΛΑ ",  font=("arial black", 11), command=open_total,  bd=6, background="#06D784", activebackground="#06D784", activeforeground="#06D784")
    total_btn.grid(row=12, columnspan=2, pady=(20, 5), ipadx=65)




    name_label = Label(root, text="Created and Designed by : Papaioannou Antonios", font=("arial black", 10), foreground="grey", background="#EEE8D1", borderwidth=1)
    name_label.grid(column=0, row=13, sticky=E, pady=(5, 0))

    select_btn = Button(root, text="Επιλογή Ποικιλίας :", command=select_variety, font=("arial black", 11), background="#0FAFC3")
    select_btn.grid(row=9, column=0, sticky=W, padx=50, pady=10)







    conn.commit()


    root.mainloop()